#!/usr/local/bin python3

from shutil import copyfile, move, copy2, copyfileobj
from openpyxl import load_workbook, Workbook, utils
from tqdm import tqdm

import zipfile2, csv, re, fnmatch, os, sys, time, datetime, random, statistics, inspect
from docx import Document

"""
OSx:
PATH="$PATH:/Library/Frameworks/Python.framework/Versions/3.5/bin/"
pyinstaller -F script_osx_remove_author.py
pyinstaller -F --additional-hooks-dir='.' script.py

WIN:
pyinstaller -F script.py

"""

staff = []
assesment_for_each_staff = []
folders_to_zip = []
num_assesmentfolder = 0
path = os.path.dirname(os.path.realpath(sys.argv[0]))
log_fil = open(path+"/script_Log.log", "w")
log_fil.close()
distribution_grade = dict()
grade_value = []
feedback_files_docx = []
feedback_files_pdf = []
assessment_criteria = [['Introduction',20], ['Main part',40], ['Conclusion',20], ['Reference',20]]
match = '*_assign*'

def log(whoami,msg):
    log_fil = open(path+"/script_Log.log", "a")
    ts = time.time()
    st = datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S')
    log_fil.write(st + ':\t'+whoami+'-->\t' + msg + '\n')
    log_fil.close()

def whoami():
    return inspect.stack()[1][3]

def zipdir(path_zip, ziph):
    log(whoami(), 'path: {}'.format(path_zip))
    log(whoami(), 'full path: {}'.format(os.path.join(path, path_zip)))

    for root, dirs, files in os.walk(os.path.join(path, path_zip)):
        for file in files:
            log(whoami(), 'Writing zip {}'.format(os.path.join(root, file),os.path.relpath(os.path.join(root, file), os.path.join(path, '..'))))
            ziph.write(os.path.join(root, file), os.path.relpath(os.path.join(root, file), os.path.join(path, '..')))

def make_zip_to_staff(path):
    log(whoami(), 'folders_to_zip: {} {}'.format(folders_to_zip, len(folders_to_zip)))
    log(whoami(), 'staff: {} {}'.format(staff, len(staff)))
    pbar = tqdm(total=len(folders_to_zip))

    for i in range(len(staff)):
        zf = zipfile2.ZipFile(path + '/' + staff[i] +'.zip', 'w', zipfile2.ZIP_DEFLATED)
        log(whoami(), 'folders_to_zip: {}'.format(folders_to_zip[i]))
        for j in range(len(folders_to_zip[i])):
            pbar.update(1)
            time.sleep(1)
            try:
                zipdir(folders_to_zip[i][j] + '/', zf)
            finally:
                pass
        log(whoami(),'Created: '+staff[i] + '.zip')
        log(whoami(),'\t\tinfo: {}'.format(zf.infolist()))
        zf.close()
        pbar.close()

def search_dir(path,*args, **kwargs):
    end_sufix = args
    log(whoami(), 'end_sufix: {}'.format(end_sufix))
    global file_list
    file_list = []
    files = os.listdir(path=path)
    for docxname in files:
        if docxname.endswith(end_sufix[0]):
            file_list.append(os.path.join(path, docxname))
    return file_list

def print_dir():
    print("Path: " + path)
    i = 0
    for docxname in file_list:
        print(i, ': ' + docxname)
        i += 1

def distribute_number_of_exam():
    number_of_staff = len(staff)
    number_for_each_staff = (num_assesmentfolder / number_of_staff)
    remaining_assesments = (num_assesmentfolder % number_of_staff)

    for i in range(1, number_of_staff + 1):
        print(i , ' - ', staff[i-1])
        extra = 1 if i <= remaining_assesments else 0
        assesment_for_each_staff.append(int(number_for_each_staff + extra))

    log(whoami(),'Distribution: '+str(assesment_for_each_staff))
    log(whoami(),'Sum of Distribution: '+str(sum(assesment_for_each_staff)))

def select_staff():
    input_names = input('Type in names of who is going to assess. Use comma between if there is more than one:')
    print(input_names)
    global staff
    staff = [x.strip() for x in input_names.split(',')]
    random.shuffle(staff)
    log(whoami(),'#{} - Staff: {}'.format(len(staff), str(staff)))
    global num_assesmentfolder
    global match
    tmp_match = input('What should it look for in selection students assessments. Default=:{} '.format(match))
    if len(tmp_match) > 1:
        match = tmp_match
    print(match)
    num_assesmentfolder = count_assesment_folders(match)
    global folders_to_zip
    folders_to_zip = [[] for i in range(len(staff))]
    distribute_number_of_exam()

def count_assesment_folders(match):
    number_of_assesments = 0
    files = os.listdir(path=path)
    for name in files:
        if fnmatch.fnmatch(name, match):
            log(whoami(),'FNMatch '+ name)
            number_of_assesments += 1
    log(whoami(), 'Number of assesments: ' + str(number_of_assesments))
    return number_of_assesments

def make_dir_to_upload():
    date = input('Type in the due date, YYYY-MM-DD >')
    course = input('Type in the course abbreviation >')
    startup = input('Type in the course startup ex:AUG-FT-2018 >')

    dirname = date+' - '+course+' - '+startup

    if not os.path.isdir(os.path.join(path, dirname)):
        os.mkdir(os.path.join(path, dirname))

    if not os.path.isdir(os.path.join(path, dirname, 'completed')):
        os.mkdir(os.path.join(path, dirname, 'completed'))

    return os.path.join(path, dirname)

def create_sheet_header_info(wb):
    ws = wb.create_sheet("Distribution", 0)
    ws.cell(row=1, column=1, value='ID')
    ws.cell(row=1, column=2, value='Teacher')
    ws.cell(row=1, column=3, value='Start')
    ws.cell(row=1, column=4, value='Finished')
    ws.cell(row=1, column=5, value='Points')
    ws.cell(row=1, column=6, value='Grade')
    c=7
    for criteria in assessment_criteria:
        ws.cell(row=1, column=c, value=criteria[0])
        c+=1
    ws.cell(row=1, column=c, value='Comments')
    c=c+5
    ws.cell(row=1, column=c+1, value='Teacher')
    ws.cell(row=1, column=c+2, value='Total')
    ws.cell(row=1, column=c+3, value='Done')
    ws.cell(row=1, column=c+4, value='Left')
    total = count_assesment_folders(match) + 2
    for x in range(0,len(staff)):
        ws.cell(row=x+2, column=c+1, value=staff[x])
        ws.cell(row=x+2, column=c+2, value=assesment_for_each_staff[x])
        ws.cell(row=x+2, column=c+3, value='=COUNTIFS($B$2:$B${0},{1}{2},$D$2:$D${0},"x")'.format(total,utils.get_column_letter(c+1),x+2))
        ws.cell(row=x+2, column=c+4, value='={}{}-{}{}'.format(utils.get_column_letter(c+2),x+2, utils.get_column_letter(c+3),x+2))

    ws.cell(row=x+3, column=c+1, value='Total')
    ws.cell(row=x+3, column=c+2, value='=SUM({}{}:{}{})'.format(utils.get_column_letter(c+2),2, utils.get_column_letter(c+2),x+2))
    ws.cell(row=x+3, column=c+3, value='=SUM({}{}:{}{})'.format(utils.get_column_letter(c+3),2, utils.get_column_letter(c+3),x+2))
    ws.cell(row=x+3, column=c+4, value='=SUM({}{}:{}{})'.format(utils.get_column_letter(c+4),2, utils.get_column_letter(c+4),x+2))


    return ws

def create_feedbackfiles_turnitin():
    global folders_to_zip, assessment_criteria
    log(whoami(),'Create_feedbackfile')
    file_list = search_dir(path, '.docx', '.doc')
    print_dir()
    log(whoami(), str(file_list))
    feedback_file_path = file_list[int(input('Type in the number of the feedback file: '))]
    folder_name = os.path.dirname(path)
    log(whoami(),"Path: " + path)
    log(whoami(),"Folder name: " + folder_name)
    tmp = input('Do you want to change the default assessment criteria? Y/N \n {}'.format(assessment_criteria))
    if tmp.lower() == 'y':
        assessment_criteria.clear()
        temp = ''
        while True:
            temp = input('Name of the criteria, type in done to end!:')
            if temp == 'done' or temp == 'Done':
                break

            temp_score = int(input('Total score on criteria:'))
            assessment_criteria.append([temp,temp_score])
    path_upload = make_dir_to_upload()
    wb = Workbook()
    ws = create_sheet_header_info(wb)
    log(whoami(),'Created Headers to Distribution.xlsx')
    i = 0
    x = 0
    z = 1
    for dirname, dirnames, filenames in os.walk(path):
        # print path to all subdirectories first.
        for name in filenames:
            if fnmatch.fnmatch(name, match):
                name = name.split(".pdf")[0]
                log(whoami(),'subdirname: ' + name)
                studentID = re.findall(r'\d+', name)
                log(whoami(),'len(studentID):' + str(len(studentID)))
                log(whoami(),'assessment_for_each_staff {} '.format(assesment_for_each_staff))
                log(whoami(),'x: {} -- i: {}'.format(x, i))
                log(whoami(),'Staff: {}'.format(staff))
                if x >= assesment_for_each_staff[i]:
                    i += 1
                    x = 0
                print(i)
                print(staff)
                staff_name = staff[i]
                x += 1
                if len(studentID) == 1:
                    z += 1
                    log(whoami(),'studentID: {} {}'.format(studentID[0], type(studentID[0])))
                    ws.cell(row=z, column=1, value=studentID[0])
                    ws.cell(row=z, column=2, value=staff_name)

                    if len(assessment_criteria) == 4:
                        ws.cell(row=z, column=5, value='={ac[0][1]}*{cell1}% + {ac[1][1]}*{cell2}% + {ac[2][1]}*{cell3}% + {ac[3][1]}*{cell4}%'.format(
                            ac=assessment_criteria,
                            cell1=utils.get_column_letter(7)+str(z),
                            cell2=utils.get_column_letter(8)+str(z),
                            cell3=utils.get_column_letter(9)+str(z),
                            cell4=utils.get_column_letter(10)+str(z)
                        ))
                        ws.cell(row=z, column=6, value='=IF({0}{1}<40,"F",IF({0}{1}<50,"E",IF({0}{1}<60,"D",IF({0}{1}<80,"C",IF({0}{1}<90,"B","A")))))'.format(utils.get_column_letter(5),z))

                    elif len(assessment_criteria) == 5:
                        ws.cell(row=z, column=5, value='={ac[0][1]}*{cell1}% + {ac[1][1]}*{cell2}% + {ac[2][1]}*{cell3}% + {ac[3][1]}*{cell4}% + {ac[4][1]}*{cell5}%'.format(
                            ac=assessment_criteria,
                            cell1=utils.get_column_letter(7)+str(z),
                            cell2=utils.get_column_letter(8)+str(z),
                            cell3=utils.get_column_letter(9)+str(z),
                            cell4=utils.get_column_letter(10)+str(z),
                            cell5 = utils.get_column_letter(11) + str(z)
                        ))
                        ws.cell(row=z, column=6, value='=IF({0}{1}<40,"F",IF({0}{1}<50,"E",IF({0}{1}<60,"D",IF({0}{1}<80,"C",IF({0}{1}<90,"B","A")))))'.format(utils.get_column_letter(5),z))

                    else:
                        print('The number of assessment criteria is not correct!')

                    wb.save(path_upload+'/Distribution.xlsx')
                print('NAME-Folder to zip', name)
                folders_to_zip[i].append(name)
                os.makedirs(path + '/' + name)
                copyfile(path+ '/' +name+'.pdf',path + '/' + name+ '/'+name+'.pdf')
                copyfile(feedback_file_path, path + '/' + name + '/' + name + '.docx')
                os.remove(path+ '/' +name+'.pdf')
                log(whoami(),'Made new file: ' + path + '/' + name + '/' + name + '.docx')

    log(whoami(),'Saved '+path+'/Distribution.xlsx')
    log(whoami(),'Folders to zip (id_staff): ' + str(folders_to_zip))

    make_zip_to_staff(path_upload)
    if not os.path.isfile(os.path.join(path, 'completed')):
        os.mkdir(os.path.join(path, 'completed'))

    print(
        '\nProcess done!\nYou will find a document in each folder and a Distribution.xlsx with all student identifikation numbers')
    log(whoami(),
        'Process done! You will find a document in each folder and a Distribution.xlsx with all student identifikation numbers')

def create_feedbackfiles():
    global folders_to_zip, assessment_criteria
    log(whoami(),'Create_feedbackfile')
    file_list = search_dir(path, '.docx', '.doc')
    print_dir()
    log(whoami(), str(file_list))
    feedback_file_path = file_list[int(input('Type in the number of the feedback file: '))]
    folder_name = os.path.dirname(path)
    log(whoami(),"Path: " + path)
    log(whoami(),"Folder name: " + folder_name)
    tmp = input('Do you want to change the default assessment criteria? Y/N \n {}'.format(assessment_criteria))
    if tmp.lower() == 'y':
        assessment_criteria.clear()
        temp = ''
        while True:
            temp = input('Name of the criteria, type in done to end!:')
            if temp == 'done' or temp == 'Done':
                break
            temp_score = int(input('Total score on criteria:'))
            assessment_criteria.append([temp,temp_score])
    path_upload = make_dir_to_upload()
    wb = Workbook()
    ws = create_sheet_header_info(wb)
    log(whoami(),'Created Headers to Distribution.xlsx')

    i = 0
    x = 0
    z = 1
    for dirname, dirnames, filenames in os.walk(path):
        for subdirname in dirnames:
            if fnmatch.fnmatch(subdirname, match):
                log(whoami(),'subdirname: ' + subdirname)
                studentID = re.findall(r'\d+', subdirname)
                log(whoami(),'len(studentID):' + str(len(studentID)))
                log(whoami(),'assessment_for_each_staff {} '.format(assesment_for_each_staff))
                log(whoami(),'x: {} -- i: {}'.format(x, i))
                log(whoami(),'Staff: {}'.format(staff))
                if x >= assesment_for_each_staff[i]:
                    i += 1
                    x = 0
                staff_name = staff[i]
                x += 1
                if len(studentID) == 1:
                    z += 1
                    log(whoami(),'studentID: {} {}'.format(studentID[0], type(studentID[0])))
                    ws.cell(row=z, column=1, value=studentID[0])
                    ws.cell(row=z, column=2, value=staff_name)
                    if len(assessment_criteria) == 4:
                        ws.cell(row=z, column=5, value='={ac[0][1]}*{cell1}% + {ac[1][1]}*{cell2}% + {ac[2][1]}*{cell3}% + {ac[3][1]}*{cell4}%'.format(
                            ac=assessment_criteria,
                            cell1=utils.get_column_letter(7)+str(z),
                            cell2=utils.get_column_letter(8)+str(z),
                            cell3=utils.get_column_letter(9)+str(z),
                            cell4=utils.get_column_letter(10)+str(z)
                        ))
                        ws.cell(row=z, column=6, value='=IF({0}{1}<40,"F",IF({0}{1}<50,"E",IF({0}{1}<60,"D",IF({0}{1}<80,"C",IF({0}{1}<90,"B","A")))))'.format(utils.get_column_letter(5),z))

                    wb.save(path_upload+'/Distribution.xlsx')
                folders_to_zip[i].append(subdirname)
                copyfile(feedback_file_path, path + '/' + subdirname + '/' + subdirname + '.docx')
                log(whoami(),'Made new file: ' + path + '/' + subdirname + '/' + subdirname + '.docx')

    log(whoami(),'Saved '+path_upload+'/Distribution.xlsx')
    log(whoami(),'Folders to zip (id_staff): ' + str(folders_to_zip))

    make_zip_to_staff(path_upload)


    print(
        '\nProcess done!\nYou will find a document in each folder and a Distribution.xlsx with all student identifikation numbers')
    log(whoami(),
        'Process done! You will find a document in each folder and a Distribution.xlsx with all student identifikation numbers')

def merge_csv_sheet():
    read_xlsx_file()
    log(whoami(),'Done read_xlsx_file')
    read_csv_file()
    log(whoami(),'Done read_csv_file')
    print(
        '\nProcess done!\nYou will find a NEW csv file ready to upload to moodle')
    log(whoami(),
        'Process done! You will find a NEW csv file ready to upload to moodle')

def copyLargeFile(src, dest, buffer_size=16000):
    with open(src, 'rb') as fsrc:
        with open(dest, 'wb') as fdest:
            copyfileobj(fsrc, fdest, buffer_size)

def move_student_exam():
    log(whoami(),'Move student exam start!')
    print(os.path.isdir(os.path.join(path, 'pdf')))
    if not os.path.isdir(os.path.join(path, 'pdf')):
        os.mkdir(os.path.join(path, 'pdf'))
    for dirname, dirnames, filenames in os.walk(path):
        # print path to all subdirectories first.
        for subdirname in dirnames:
            if fnmatch.fnmatch(subdirname, '*_assign*'):
                for subsubdirname, subdirnames, subdirfiles in os.walk(os.path.join(path,subdirname)):
                    for subdirfile in subdirfiles:
                        if subdirfile.endswith("pdf"):
                            newfilename = "{}{}".format(subdirname, subdirfile)
                            log(whoami(), "{}{}".format(os.path.join(dirname, subdirname, subdirfile) ,os.path.join(path, 'pdf', newfilename)))
                            copyLargeFile(os.path.join(dirname, subdirname), os.path.join(path, 'pdf', newfilename))

def delete_student_exam():
    for dirname, dirnames, filenames in os.walk(path):
        # print path to all subdirectories first.
        for docxname in filenames:
            print(dirname)
            print(': ' + docxname)
            if not fnmatch.fnmatch(docxname, 'FS*'):
                if not fnmatch.fnmatch(docxname, '*py'):
                    print('File: ' + docxname)
                    os.remove(dirname + '/' + docxname)

def calculate_stats():
    print("STAT:", type(grade_value), grade_value)
    try:
        stat_file = open(path+'/CP_statistics.txt', 'w')
        stat_file.write('Mean:\t {} - Arithmetic mean (“average”) of data.\n'.format(statistics.mean(grade_value)))
        stat_file.write('Median:\t {} - Median (middle value) of data.\n'.format(statistics.median(grade_value)))
        stat_file.write('Median_low:\t {} - Low median of data.\n'.format(statistics.median_low(grade_value)))
        stat_file.write('Median_high:\t {} - High median of data.\n'.format(statistics.median_low(grade_value)))
        stat_file.write('Median_grouped:\t {} - Median, or 50th percentile, of grouped data.\n'.format(statistics.median_grouped(grade_value)))
        stat_file.write('Population standard deviation:\t {} - Population standard deviation of data.\n'.format(statistics.pstdev(grade_value)))
        stat_file.write('Population variance:\t {} - Population variance of data.\n'.format(statistics.pvariance(grade_value)))
        stat_file.write('Standard deviation:\t {} - Sample standard deviation of data.\n'.format(statistics.stdev(grade_value)))
        stat_file.write('Variance:\t {} - Sample variance of data.\n'.format(statistics.variance(grade_value)))
        stat_file.close()
        log(whoami(),'Done calculated stats')
    except TypeError as msg:
        log(whoami(), msg)

def read_xlsx_file():
    log(whoami(),'read_xlsx_file')
    file_list = search_dir(path, '.xlsx')
    print_dir()
    dist_xlsx_path = file_list[int(input('Type in the number of the Distribution sheet:'))]
    log(whoami(),'dist_xlsx_path -> '+ dist_xlsx_path)
    wb = load_workbook(filename=dist_xlsx_path, read_only=True, data_only=True)
    ws = wb['Distribution']
    for row in ws.rows:
        if row[0].value == 'ID':
            continue
        global distribution_grade
        distribution_grade[row[0].value] = []
        distribution_grade[row[0].value].append([row[4].value, row[5].value])
        global grade_value
        grade_value.append(row[4].value)

def read_csv_file():
    log(whoami(),'read_csv_file')
    file_list = search_dir(path, '.csv')
    print_dir()
    dist_csv_path = file_list[int(input('Type in the number of the Grade sheet from Moodle sheet:'))]
    log(whoami(),'dist_csv_path -> '+ dist_csv_path)
    with open(dist_csv_path, "r", newline='', encoding='utf-8') as csv_file:
        reader = csv.DictReader(csv_file, delimiter=',')
        log(whoami(),'distribution_grade: {} '.format(distribution_grade))
        with open(path+'/NEW-Greeding-upload.csv', 'w', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, delimiter=',', fieldnames=reader.fieldnames)
            writer.writeheader()
            for row in reader:
                p_id = re.findall(r'\d+', row['\ufeffIdentifier'])[0]
                log(whoami(),'p_id: {} {} in: {} '.format(p_id,type(p_id), p_id in distribution_grade))
                #p_id = int(p_id)
                #log(whoami(),'p_id: {} {} in: {} '.format(p_id,type(p_id), p_id in distribution_grade))
                log(whoami(), 'p_id in dist: {}'.format(p_id in distribution_grade))
                if p_id in distribution_grade:
                    writer.writerow({'\ufeffIdentifier': row['\ufeffIdentifier'],
                                     'Status': row['Status'],
                                     'Grade': distribution_grade[p_id][0][0],
                                     #'Scale': row['Scale'],
                                     'Maximum Grade': row['Maximum Grade'],
                                     'Grade can be changed': row['Grade can be changed'],
                                     'Last modified (submission)': row['Last modified (submission)'],
                                     'Last modified (grade)': row['Last modified (grade)'],
                                     'Feedback comments': 'Grade = {}'.format(distribution_grade[p_id][0][1])})
                                     #'Feedback comments': 'Points = {}'.format(distribution_grade[p_id][0][0])})

def collect_feedback_files(completed_file_path):
    log(whoami(), 'Collect Feedback files')
    for dirname, dirnames, filenames in os.walk(completed_file_path):
        global feedback_files_docx
        for docname in filenames:
            if docname[-5:] == ".docx":
                path = os.path.abspath(os.path.join(completed_file_path, dirname, docname))
                feedback_files_docx.append(path)
        change_docx_attributes()

def change_docx_attributes():
    for docname in feedback_files_docx:
        log(whoami(), 'Docname: {}'.format(docname))
        doc = Document(docname)
        core_properties = doc.core_properties
        meta_fields= ["author", "last_modified_by"]
        for meta_field in meta_fields:
            setattr(core_properties, meta_field, "Noroff-IT Team")
            log(whoami(), 'Author change to {}'.format("Noroff-IT Team"))
        doc.save(docname)

#Preparing for making feedback files.
def make_feedback_zip():
    file_list = search_dir(path, 'completed')
    print_dir()
    completed_file_path = file_list[int(input('Type in the number of the feedback file folder: '))]
    log(whoami(), 'MAKEDIR')
    log(whoami(), 'MAKEDIR PATH {}'.format(completed_file_path))
    makedir(completed_file_path)

#Creating folder structure for ZIP and create feedback.zip
def makedir(completed_file_path):
    log(whoami(), 'MAKEDIR START - Creating folder structure to feedback.zip')
    files = os.listdir(path=completed_file_path)
    for docxname in files:
        #log(whoami(), 'MAKEDIR FOR: {} {}'.format(completed_file_path, docxname))
        try:
            #log(whoami(),'File exist: {}'.format(os.path.isfile(os.path.join(completed_file_path.decode(), docxname.strip('.docx')))))
            if not os.path.isfile(os.path.join(completed_file_path, docxname.strip('.docx'))):
                os.mkdir(os.path.join(completed_file_path, docxname.strip('.docx')))
                if os.path.isfile(os.path.join(completed_file_path, docxname)):
                    move(os.path.join(completed_file_path, docxname), os.path.join(completed_file_path, docxname.strip('.docx')))
        except FileExistsError as msg:
            log(whoami(),'{}'.format(msg))

    collect_feedback_files(completed_file_path)
    zf = zipfile2.ZipFile(os.path.join(path,'Feedback.zip'), 'w', zipfile2.ZIP_DEFLATED)
    log(whoami(),'ZIP: {}'.format(os.path.join(path,'Feedback.zip')))
    folders_to_zip_feedback = search_dir(os.path.join(path, 'completed/'),'_file_')
    log(whoami(), 'folders_to_zip_feedback: {}'.format(folders_to_zip_feedback))
    for folder_path in folders_to_zip_feedback:
        try:
            zipdir(folder_path, zf)
        finally:
            log(whoami(), 'ZIP: {}'.format(zf.infolist()))

def hasNumbers(inputString):
    return bool(re.search(r'\d', inputString))

def check_missing_feedback():
    log(whoami(), 'Start missing feedback files')
    for dirname, dirnames, filenames in os.walk(path):
       for dir in dirnames:
           if dir == 'completed':
            feedback_path = os.path.join(path, dir)
    id_list = set()
    for dirname, dirnames, filenames in os.walk(feedback_path):
       for file in filenames:
           if hasNumbers(file):
            id_list.add(re.findall('\d+', file )[0])

    dist_id_name = dict()
    file_list = search_dir(path, '.xlsx')
    print_dir()
    dist_xlsx_path = file_list[int(input('Type in the number of the Distribution sheet:'))]
    log(whoami(), 'dist_xlsx_path -> ' + dist_xlsx_path)
    wb = load_workbook(filename=dist_xlsx_path, read_only=True, data_only=True)
    ws = wb['Distribution']
    for row in ws.rows:
        if row[0].value == 'ID':
            continue
        dist_id_name[row[0].value] = row[1].value
    keys_diff = set(dist_id_name.keys()) - id_list
    missing_feedback_log = open(path + "/missing_feedback.txt", "w")
    for key in keys_diff:
        missing_feedback_log.write(str(key) + ' -- '+ str(dist_id_name[key])+'\n')
    missing_feedback_log.close()



log_fil = open(path+"/script_Log.log", "w")
log_fil.write('LOG FOR ASSESSMENT SCRIPT\n')
log_fil.close()

prog_to_run = -1
while prog_to_run != 0:
    prog_to_run = int(input('What program/operation do you want to run? Type in the number, 0 to quit:\n'
                        '\t1: Create feedback file in each folder, and collect the student ID in a list.\n'
                        '\t2: Merge grades into feedback file with merge dist.list and Moodle grade sheet.\n'
                        '\t3: Make feedback zip.\n'
                        '\t4: Check missing feedback files.\n'))

    if prog_to_run == 1:
        select_staff()
        tmp = input('Do you use TurnItIn? Yes/No')
        if tmp.lower() == "yes":
            print("using TurnItIn")
            create_feedbackfiles_turnitin()
        else:
            print('Normal run')
            create_feedbackfiles()
    elif prog_to_run == 2:
        merge_csv_sheet()
    elif prog_to_run == 3:
        make_feedback_zip()
    elif prog_to_run == 4:
        check_missing_feedback()
    elif prog_to_run == 0:
        sys.exit(0)
    else:
        print('Type in one of the number to choose select a script_osx_remove_author!')
